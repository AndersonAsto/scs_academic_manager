import mysql.connector
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import itertools

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "sv_academic_manager",
}

DIAS = [
    "Lunes",
    "Martes",
    "Miércoles",
    "Jueves",
    "Viernes"
]

COURSE_COLORS = [
    "FFD966",
    "9BC2E6",
    "A9D18E",
    "F4B084",
    "C6E0B4",
    "D9D2E9",
    "FCE4D6",
    "BDD7EE",
    "E2EFDA",
    "FFF2CC"
]

# ----------------------------------------------------
# LOAD DATA
# ----------------------------------------------------

conn = mysql.connector.connect(**DB_CONFIG)
cursor = conn.cursor(dictionary=True)

cursor.execute("""
SELECT

    y.year,

    g.grade,
    sec.section,

    c.course,

    ts.time_slot,
    ts.start_time,

    s.day,

    CONCAT(
        pi.names,' ',
        pi.fathers_surname,' ',
        pi.mothers_surname
    ) AS teacher

FROM schedules s

INNER JOIN teacher_groups tg
    ON tg.id = s.teacher_group_id

INNER JOIN academic_staff_contracts ascn
    ON ascn.id = tg.academic_staff_contract_id

INNER JOIN academic_staff ast
    ON ast.id = ascn.academic_staff_id

INNER JOIN personal_information pi
    ON pi.id = ast.personal_information_id

INNER JOIN years y
    ON y.id = ascn.year_id

INNER JOIN courses c
    ON c.id = tg.course_id

INNER JOIN grades g
    ON g.id = tg.grade_id

INNER JOIN sections sec
    ON sec.id = tg.section_id

INNER JOIN time_slots ts
    ON ts.id = s.time_slot_id

WHERE
    s.status = 1
    AND tg.status = 1
    AND ascn.status = 1
    AND ast.status = 1

ORDER BY
    y.year,
    g.id,
    sec.id,
    ts.start_time
""")

rows = cursor.fetchall()

cursor.close()
conn.close()

# ----------------------------------------------------
# ORGANIZAR
# ----------------------------------------------------

data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))

courses = set()

time_slots = []

for r in rows:

    year = r["year"]

    group = f"{r['grade']} {r['section']}"

    slot = r["time_slot"]

    if slot not in time_slots:
        time_slots.append(slot)

    data[year][group][slot][r["day"]] = (
        r["course"],
        r["teacher"]
    )

    courses.add(r["course"])

# ----------------------------------------------------
# COLORES
# ----------------------------------------------------

course_color = {}

for c, color in zip(sorted(courses), itertools.cycle(COURSE_COLORS)):
    course_color[c] = PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid"
    )

thin = Side(style="thin")

# ----------------------------------------------------
# EXCEL
# ----------------------------------------------------

wb = Workbook()
wb.remove(wb.active)

for year, groups in data.items():

    ws = wb.create_sheet(title=f"{year}")

    row = 1

    ws.cell(
        row=row,
        column=1,
        value=f"HORARIOS - AÑO {year}"
    ).font = Font(size=15, bold=True)

    row += 2

    for group in sorted(groups):

        ws.cell(
            row=row,
            column=1,
            value=f"GRADO / SECCIÓN : {group}"
        ).font = Font(size=12, bold=True)

        row += 1

        # encabezado

        ws.cell(
            row=row,
            column=1,
            value="Horario"
        ).font = Font(bold=True)

        for col, d in enumerate(DIAS, start=2):

            c = ws.cell(
                row=row,
                column=col,
                value=d
            )

            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        row += 1

        # filas

        for slot in time_slots:

            ws.cell(
                row=row,
                column=1,
                value=slot
            )

            for col, day in enumerate(DIAS, start=2):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                if day in groups[group][slot]:

                    course, teacher = groups[group][slot][day]

                    cell.value = f"{course}\n\n{teacher}"

                    cell.fill = course_color[course]

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

            ws.row_dimensions[row].height = 45

            row += 1

        row += 2

    # ancho columnas

    ws.column_dimensions["A"].width = 22

    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 30

# ----------------------------------------------------
# GUARDAR
# ----------------------------------------------------

filename = (
    f"Horarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

wb.save(filename)

print(f"Excel generado correctamente: {filename}")