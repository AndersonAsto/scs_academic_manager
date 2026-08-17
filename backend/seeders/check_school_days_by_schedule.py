"""
Genera un libro de Excel por cada grado-sección (1A, 1B, 2A ... 5B) con el
calendario completo de días lectivos: fecha, día de la semana, curso, hora y
tipo de evaluación (Calificación Diaria / Práctica / Examen).

Sirve para verificar visualmente que la asignación de prácticas y exámenes
generada por seed_school_days_by_schedule.py sea correcta.

Requiere: mysql-connector-python, openpyxl
"""

import os
import re
import mysql.connector
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "sv_academic_manager",
    "charset": "utf8mb4",
}

OUTPUT_DIR = "output/calendars"

# Colores por tipo de evaluación (fondo)
COLOR_POR_TIPO = {
    "Calificación Diaria": "D9EAD3",  # verde claro
    "Práctica": "FFF2CC",             # amarillo claro
    "Examen": "F4CCCC",               # rojo claro
}

FUENTE = "Arial"


# -----------------------------
# 1. CONSULTA
# -----------------------------
def obtener_datos():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            g.grade          AS grado,
            se.section       AS seccion,
            c.course         AS curso,
            s.day            AS dia_semana,
            ts.time_slot     AS hora,
            sd.school_day    AS fecha,
            sd.week_number   AS semana,
            tb.teaching_block AS bloque,
            sds.type         AS tipo
        FROM teacher_groups tg
        JOIN grades g            ON g.id = tg.grade_id
        JOIN sections se         ON se.id = tg.section_id
        JOIN courses c           ON c.id = tg.course_id
        JOIN schedules s         ON s.teacher_group_id = tg.id AND s.status = 1
        JOIN school_days_by_schedule sds ON sds.schedule_id = s.id AND sds.status = 1
        JOIN school_days sd      ON sd.id = sds.school_day_id AND sd.status = 1
        JOIN teaching_blocks tb  ON tb.id = sd.teaching_block_id
        JOIN time_slots ts       ON ts.id = s.time_slot_id
        WHERE tg.status = 1
        ORDER BY g.grade, se.section, sd.school_day, ts.start_time
    """)
    filas = cursor.fetchall()

    cursor.close()
    conn.close()
    return filas


# -----------------------------
# 2. AGRUPAR POR GRADO-SECCION
# -----------------------------
def agrupar_por_grado_seccion(filas):
    grupos = defaultdict(list)
    for f in filas:
        clave = (f['grado'], f['seccion'])
        grupos[clave].append(f)
    return grupos


def nombre_archivo(grado, seccion):
    base = f"{grado}{seccion}"
    # quita caracteres no válidos para nombre de archivo (° espacios etc.)
    base = re.sub(r'[^A-Za-z0-9]+', '', base)
    return f"{base}.xlsx"


# -----------------------------
# 3. CONSTRUCCIÓN DEL EXCEL
# -----------------------------
def construir_libro(grado, seccion, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"

    encabezados = ["Bloque", "Semana", "Fecha", "Día", "Curso", "Hora", "Tipo"]

    fuente_header = Font(name=FUENTE, bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="4472C4")
    borde = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
    alineacion = Alignment(horizontal="center", vertical="center")

    ws.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = fuente_header
        celda.fill = fill_header
        celda.alignment = alineacion
        celda.border = borde

    fila_actual = 2
    for f in filas:
        valores = [
            f['bloque'],
            f['semana'],
            f['fecha'].strftime("%d/%m/%Y"),
            f['dia_semana'],
            f['curso'],
            f['hora'],
            f['tipo'] or "Calificación Diaria",
        ]
        ws.append(valores)

        color = COLOR_POR_TIPO.get(f['tipo'], COLOR_POR_TIPO["Calificación Diaria"])
        fill_tipo = PatternFill("solid", fgColor=color)

        for col_idx in range(1, len(encabezados) + 1):
            celda = ws.cell(row=fila_actual, column=col_idx)
            celda.font = Font(name=FUENTE, size=10)
            celda.border = borde
            celda.alignment = alineacion if col_idx != 5 else Alignment(horizontal="left", vertical="center")

        # solo la columna "Tipo" (7) recibe el color distintivo
        ws.cell(row=fila_actual, column=7).fill = fill_tipo

        fila_actual += 1

    # anchos de columna
    anchos = [12, 9, 12, 12, 30, 14, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{fila_actual - 1}"

    return wb


# -----------------------------
# 4. MAIN
# -----------------------------
def main():
    filas = obtener_datos()
    print(f"[debug] filas obtenidas: {len(filas)}")

    if not filas:
        print("No se encontraron registros en school_days_by_schedule. "
              "Corre primero el seeder de evaluaciones.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    grupos = agrupar_por_grado_seccion(filas)
    print(f"[debug] grados-secciones encontrados: {len(grupos)}")

    for (grado, seccion), filas_grupo in grupos.items():
        wb = construir_libro(grado, seccion, filas_grupo)
        nombre = nombre_archivo(grado, seccion)
        ruta = os.path.join(OUTPUT_DIR, nombre)
        wb.save(ruta)
        print(f"  Generado: {ruta} ({len(filas_grupo)} filas)")

    print("Exportación completada.")


if __name__ == "__main__":
    main()